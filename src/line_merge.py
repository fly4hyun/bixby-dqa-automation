###################################################################################################
###################################################################################################

import os
import argparse
import openpyxl

###################################################################################################
###################################################################################################

def parse_opt():
    parser = argparse.ArgumentParser()
    
    ### default='pdf_parser_results'
    parser.add_argument('--xlsx_path', type=str, default='pdf_parser_results')
    
    ### 0: 첫번째 시트, 1: 두번째 시트
    parser.add_argument('--sheet_num', type=int, default=0)

    opt = parser.parse_args()
    return opt

###################################################################################################
###################################################################################################
### 대제목, 중제목, 소제목 처리 알고리즘, 내용 처리 알고리즘

def process_file(filepath, sheet_num):
    wb = openpyxl.load_workbook(filepath)
    
    ### 0: 첫번째 시트, 1: 두번째 시트
    second_sheet_name = wb.sheetnames[sheet_num]
    ws = wb[second_sheet_name]

    # g, h, i 열
    for row in ws.iter_rows(min_row=2):
        for cell in (row[7], row[8], row[9]):
            if cell.value:
                cell.value = str(cell.value).replace('\n', ' ')
                
    # j 열
    for row in ws.iter_rows(min_row=2):
        cell = row[10]
        if cell.value:
            cell_str = str(cell.value)
            cell_str_list = cell_str.split('\n')
            temp = ''
            for one_str in cell_str_list:
                if len(one_str) == 0:
                    continue
                if one_str[-1] in ['.', '?', '!', ')', ':']:
                    temp += one_str + '\n'
                else:
                    temp += one_str + ' '
            if temp[-1] == '\n':
                temp = temp[:-1]
            cell.value = temp

    dirname, filename = os.path.split(filepath)
    name, ext = os.path.splitext(filename)
    new_filename = f"{name}{ext}"
    #new_filepath = os.path.join(dirname, new_filename)
    os.makedirs(f'pdf_parser_results', exist_ok=True)
    new_filepath = os.path.join('pdf_parser_results', new_filename)
    wb.save(new_filepath)

###################################################################################################

def main(opt):
    opt = parse_opt()
    folder = opt.xlsx_path

    if not os.path.isdir(folder):
        return

    for filename in os.listdir(folder):
        if filename.lower().endswith('.xlsx'):
            filepath = os.path.join(folder, filename)
            process_file(filepath, opt.sheet_num)

###################################################################################################
###################################################################################################

if __name__ == "__main__":
    opt = parse_opt()
    main(opt)

###################################################################################################
###################################################################################################